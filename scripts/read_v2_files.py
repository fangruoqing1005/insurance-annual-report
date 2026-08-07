# -*- coding: utf-8 -*-
"""读取 v2 表格清单 和 8家公司数据库 Excel，输出JSON文件（含标黄标记）"""
import openpyxl
import json

def collect_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        row_data = []
        for cell in row:
            val = cell.value
            is_yellow = False
            try:
                if cell.fill and cell.fill.fgColor and cell.fill.patternType:
                    rgb = cell.fill.fgColor.rgb
                    if rgb and isinstance(rgb, str) and len(rgb) >= 6:
                        hex_rgb = rgb[-6:].upper()
                        if hex_rgb in ('FFC000', 'FFFF00', 'FFD966', 'FFF2CC', 'FFEB9C', 'FFFF99', 'FFE699', 'FFCC00'):
                            is_yellow = True
            except Exception:
                pass
            row_data.append({
                'v': '' if val is None else str(val),
                'yellow': is_yellow
            })
        rows.append(row_data)
    return {'sheet': sheet_name, 'rows': rows}

out = {}
for path, key in [
    (r"C:/Users/Lenovo/Desktop/AI/年报/组件输入/表格清单-v2.xlsx", "config_v2"),
    (r"C:/Users/Lenovo/Desktop/AI/年报/数据库/8家公司_指标定位表_合并版_v2.xlsx", "db_v2"),
]:
    wb = openpyxl.load_workbook(path, data_only=True)
    out[key] = {'sheets': [collect_sheet(wb, sn) for sn in wb.sheetnames]}

with open(r"C:/Users/Lenovo/WorkBuddy/2026-08-04-15-22-31/scripts/v2_output.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=1)

# 汇总黄色单元格数量
for key in out:
    for s in out[key]['sheets']:
        n_yellow_cells = sum(1 for r in s['rows'] for c in r if c['yellow'])
        n_yellow_rows = sum(1 for r in s['rows'] if any(c['yellow'] for c in r))
        print(f"{key}/{s['sheet']}: yellow_cells={n_yellow_cells}, yellow_rows={n_yellow_rows}")
