# -*- coding: utf-8 -*-
"""生成 data.js（v1.3 版）：92家公司总览 + 37家数据库 + 163行模板
数据源：
  - 92家公司分类总览_37家指标定位表汇总_v1.3.xlsx
    Sheet1「92家公司分类总览」：92家（序号|公司名称|官网地址|公司类型|准则适用|首次发布新准则财报年份）
    Sheet2「36家指标定位表汇总」：37家×163行，期间已规范化（本期/上期/本期初/本期末/上期初/上期末）
  - _template_163_full.json：163行指标模板
"""
import openpyxl
import json
import os

BASE = r'C:/Users/Lenovo/Desktop/AI/年报/数据库'
SRC = os.path.join(BASE, '92家公司分类总览_37家指标定位表汇总_v1.3.xlsx')
TPL = r'C:/Users/Lenovo/Desktop/AI/年报/抓取/_template_163_full.json'
OUT = r'C:/Users/Lenovo/WorkBuddy/2026-08-04-15-22-31/data.js'

HEADERS = ['公司类型', '公司名称', '报告期', '报表类型', '报表名称', '指标编号', '指标名称',
           '指标来源', '关键词', '期间', '计量单位-披露', '计量单位-换算', '数值-披露', '数值-换算',
           '来源表', '行序号', '列序号']

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

# ========== 1. 92家公司总览 ==========
ws1 = wb['92家公司分类总览']
overview = []   # [{seq, name, full_name, url, type, standard, first_year}]
for row in ws1.iter_rows(min_row=4, values_only=True):
    if row[1] is None:
        continue
    seq = row[0]
    full_name = str(row[1]).strip()
    url = str(row[2]).strip() if row[2] else ''
    ctype = str(row[3]).strip() if row[3] else ''
    standard = str(row[4]).strip() if row[4] else ''
    first_year = row[5]
    if isinstance(first_year, float) and first_year == int(first_year):
        first_year = int(first_year)
    overview.append({
        'seq': seq if isinstance(seq, int) else int(seq),
        'full_name': full_name,
        'url': url,
        'type': ctype,
        'standard': standard,
        'first_year': first_year if first_year not in (None, '-', '') else None
    })
print(f'92家总览: {len(overview)} 家')

# ========== 2. 37家指标定位表汇总 ==========
ws2 = wb['36家指标定位表汇总']
raw_rows = []
n_text = 0
for row in ws2.iter_rows(min_row=2, values_only=True):
    vals = list(row)
    if not any(v is not None and str(v).strip() != '' for v in vals[:6]):
        continue
    d = {}
    for i, h in enumerate(HEADERS):
        v = vals[i] if i < len(vals) else None
        if h in ('数值-披露', '数值-换算'):
            if isinstance(v, (int, float)):
                d[h] = float(v)
            elif v is None or str(v).strip() == '':
                d[h] = None
            else:
                s = str(v).strip()
                d[h] = s
                if h == '数值-换算' and not s.replace('.', '').replace('-', '').isdigit():
                    n_text += 1
        elif h in ('行序号', '列序号'):
            if isinstance(v, (int, float)) and v == int(v):
                d[h] = int(v)
            else:
                d[h] = v
        else:
            d[h] = None if v is None else str(v).strip()
    raw_rows.append(d)
print(f'RAW_DATA 行数: {len(raw_rows)}, 文本值: {n_text}')

# 公司清单 + 类型
db_companies = {}
for r in raw_rows:
    db_companies.setdefault(r['公司名称'], r['公司类型'])
print(f'数据库公司数: {len(db_companies)}')

# 期间分布
from collections import Counter
periods = Counter(r['期间'] for r in raw_rows)
print('期间分布:', dict(periods))

# ========== 3. 163行模板 ==========
with open(TPL, 'r', encoding='utf-8') as f:
    template = json.load(f)
print(f'模板行数: {len(template)}')

# ========== 4. 构建 COMPANIES_DATA（92家，6类型）==========
# 类型体系以92家总览为准：头部险企/外资公司/银行系公司/养老公司/健康险公司/中小公司
# 用户确认的类型修正（覆盖总览原始标注）
TYPE_OVERRIDE = {
    '国民养老保险股份有限公司': '银行系公司',   # 用户2026-08-12确认：国民养老改为银行系
}
TYPE_ORDER = ['头部险企', '外资公司', '银行系公司', '养老公司', '健康险公司', '中小公司']
type_rules = {t: [] for t in TYPE_ORDER}
for o in overview:
    t = TYPE_OVERRIDE.get(o['full_name'], o['type'])
    o['type'] = t
    if t in type_rules:
        type_rules[t].append(o['full_name'])
for t in TYPE_ORDER:
    print(f'  {t}: {len(type_rules[t])} 家')

# 简称映射（92家全称 → 简称，数据库公司名称可能用简称）
SHORT_NAMES = {
    '中国太平洋人寿保险股份有限公司': '太平洋人寿', '中国平安人寿保险股份有限公司': '平安人寿',
    '太平人寿保险有限公司': '太平人寿', '新华人寿保险股份有限公司': '新华人寿',
    '阳光人寿保险股份有限公司': '阳光人寿', '中国人寿保险股份有限公司': '中国人寿',
    '中国人民人寿保险股份有限公司': '人民人寿', '泰康人寿保险有限责任公司': '泰康人寿',
    '友邦人寿保险有限公司': '友邦人寿', '中意人寿保险有限公司': '中意人寿',
    '中邮人寿保险股份有限公司': '中邮人寿', '交银人寿保险有限公司': '交银人寿',
    '农银人寿保险股份有限公司': '农银人寿', '工银安盛人寿保险有限公司': '工银安盛',
    '建信人寿保险股份有限公司': '建信人寿', '中银三星人寿保险有限公司': '中银三星',
    '招商信诺人寿保险有限公司': '招商信诺', '中信保诚人寿保险有限公司': '中信保诚',
    '中荷人寿保险有限公司': '中荷人寿', '东吴人寿保险股份有限公司': '东吴人寿',
    '东方嘉富人寿保险有限公司': '东方嘉富', '国富人寿保险股份有限公司': '国富人寿',
    '太平养老保险股份有限公司': '太平养老', '太平洋健康保险股份有限公司': '太平洋健康',
    '平安健康保险股份有限公司': '平安健康', '平安养老保险股份有限公司': '平安养老',
    '中国人民健康保险股份有限公司': '中国人民健康', '中英人寿保险有限公司': '中英人寿',
    '同方全球人寿保险有限公司': '同方全球', '复星保德信人寿保险有限公司': '复星保德信',
    '恒安标准养老保险有限责任公司': '恒安标准养老', '泰康养老保险股份有限公司': '泰康养老',
    '瑞泰人寿保险有限公司': '瑞泰人寿', '财信吉祥人寿保险股份有限公司': '财信吉祥',
    '陆家嘴国泰人寿保险有限责任公司': '陆家嘴国泰', '恒安标准人寿保险有限公司': '恒安标准人寿',
    '国民养老保险股份有限公司': '国民养老',
}
# 92家总览其他公司简称（无数据库数据也补全，便于筛选显示）
EXTRA_SHORT = {
    '太平养老保险股份有限公司': '太平养老', '三峡人寿保险股份有限公司': '三峡人寿',
    '中华联合人寿保险股份有限公司': '中华联合人寿', '中宏人寿保险有限公司': '中宏人寿',
    '中美联泰大都会人寿保险有限公司': '大都会人寿', '信美人寿相互保险社': '信美相互',
    '光大永明人寿保险有限公司': '光大永明', '利安人寿保险股份有限公司': '利安人寿',
    '北京人寿保险股份有限公司': '北京人寿', '北大方正人寿保险有限公司': '北大方正',
    '华泰人寿保险股份有限公司': '华泰人寿', '华贵人寿保险股份有限公司': '华贵人寿',
    '君龙人寿保险有限公司': '君龙人寿', '和泰人寿保险股份有限公司': '和泰人寿',
    '国宝人寿保险股份有限公司': '国宝人寿', '国联人寿保险股份有限公司': '国联人寿',
    '复星联合健康保险股份有限公司': '复星联合健康', '大家养老保险股份有限公司': '大家养老',
    '安联人寿保险有限公司': '安联人寿', '小康人寿保险有限责任公司': '小康人寿',
    '幸福人寿保险股份有限公司': '幸福人寿', '弘康人寿保险股份有限公司': '弘康人寿',
    '德华安顾人寿保险有限公司': '德华安顾', '招商局仁和人寿保险股份有限公司': '招商仁和',
    '新华养老保险股份有限公司': '新华养老', '横琴人寿保险有限公司': '横琴人寿',
    '民生人寿保险股份有限公司': '民生人寿', '汇丰人寿保险有限公司': '汇丰人寿',
    '海保人寿保险股份有限公司': '海保人寿', '爱心人寿保险股份有限公司': '爱心人寿',
    '珠江人寿保险股份有限公司': '珠江人寿', '英大泰和人寿保险股份有限公司': '英大人寿',
    '长城人寿保险股份有限公司': '长城人寿', '长生人寿保险有限公司': '长生人寿',
    '中国人寿养老保险股份有限公司': '国寿养老', '中国人民养老保险有限责任公司': '人保养老',
    '长江养老保险股份有限公司': '长江养老', '富德生命人寿保险股份有限公司': '富德生命',
    '合众人寿保险股份有限公司': '合众人寿', '百年人寿保险股份有限公司': '百年人寿',
    '中汇人寿保险股份有限公司': '中汇人寿', '大家人寿保险股份有限公司': '大家人寿',
    '富泽人寿保险股份有限公司': '富泽人寿', '瑞众人寿保险有限责任公司': '瑞众人寿',
    '信泰人寿保险股份有限公司': '信泰人寿', '国华人寿保险股份有限公司': '国华人寿',
    '中融人寿保险股份有限公司': '中融人寿', '和谐健康保险股份有限公司': '和谐健康',
    '昆仑健康保险股份有限公司': '昆仑健康', '华汇人寿保险股份有限公司': '华汇人寿',
    '前海人寿保险股份有限公司': '前海人寿', '上海人寿保险股份有限公司': '上海人寿',
    '渤海人寿保险股份有限公司': '渤海人寿', '海港人寿保险股份有限公司': '海港人寿',
    '鼎诚人寿保险有限责任公司': '鼎诚人寿', '瑞华健康保险股份有限公司': '瑞华健康',
}
SHORT_NAMES.update(EXTRA_SHORT)

# 数据库公司名称 → 简称（用于 has_data 判断）
db_short_names = set(db_companies.keys())
print('数据库公司简称:', sorted(db_short_names))

# companies 数组（92家按总览顺序）
companies = []
db_used = set()
for idx, o in enumerate(overview, 1):
    short = SHORT_NAMES.get(o['full_name'], o['full_name'])
    # 数据库公司名可能等于简称（如"平安人寿"），或通过映射（"中国平安"→"平安人寿"）
    has_data = short in db_short_names
    if has_data:
        db_used.add(short)
    companies.append({
        'index': idx,
        'full_name': o['full_name'],
        'short_name': short,
        'types': [o['type']],
        'has_data': has_data,
        'url': o['url'],
        'standard': o['standard'],
        'first_year': o['first_year']
    })

print(f'COMPANIES 数量: {len(companies)}, 有数据: {sum(1 for c in companies if c["has_data"])}')
# 检查哪些数据库公司没匹配到
unmatched = db_short_names - db_used
print('未匹配到总览的数据库公司:', unmatched if unmatched else '无')

# 统一 RAW_DATA 的公司类型：以92家总览类型为准（数据库个别标注陈旧，如国民养老=银行系→养老公司）
short_to_type = {c['short_name']: c['types'][0] for c in companies}
fixed_type = 0
for r in raw_rows:
    t = short_to_type.get(r['公司名称'])
    if t and r['公司类型'] != t:
        r['公司类型'] = t
        fixed_type += 1
print(f'RAW_DATA 公司类型统一: {fixed_type} 行修正')

COMPANIES_DATA = {
    'types': TYPE_ORDER,
    'type_rules': type_rules,
    'companies': companies
}

# DB_ALIAS（77家简称 → 数据库名）：数据库公司名称以简称为主，个别映射
DB_ALIAS = {'中国平安': '平安人寿', '太平洋': '太平洋人寿', '人保': '人民人寿'}

NEGATIVE_INDICATORS = ['B07', 'B08', 'B09', 'B10', 'B12', 'D07', 'C04', 'C13']

# ========== 5. 输出 data.js ==========
lines = []
lines.append('// 数据源：92家公司分类总览_37家指标定位表汇总_v1.3.xlsx（92家总览 + 37家×163行）')
lines.append('const COMPANIES_DATA = ' + json.dumps(COMPANIES_DATA, ensure_ascii=False) + ';')
lines.append('')
lines.append('// 92家公司分类总览（展示用：序号|全称|官网|类型|准则|首年）')
lines.append('const COMPANIES_92 = ' + json.dumps(overview, ensure_ascii=False) + ';')
lines.append('')
lines.append('// 163行指标模板（智能提取页用）')
lines.append('const TEMPLATE_163 = ' + json.dumps(template, ensure_ascii=False) + ';')
lines.append('')
lines.append('// 数据库公司名称 → 77家列表简称 的别名（数据库用"平安人寿/太平洋人寿/人民人寿"）')
lines.append('const DB_ALIAS = ' + json.dumps(DB_ALIAS, ensure_ascii=False) + ';')
lines.append('')
lines.append('// 需要标准化为负值的指标（提取后统一符号）')
lines.append('const NEGATIVE_INDICATORS = ' + json.dumps(NEGATIVE_INDICATORS, ensure_ascii=False) + ';')
lines.append('')
lines.append('// 原生数据库（37家公司，6031行）')
lines.append('const RAW_DATA = ' + json.dumps(raw_rows, ensure_ascii=False) + ';')

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

size_kb = len('\n'.join(lines)) / 1024
print(f'data.js 已生成: {OUT}, 大小 {size_kb:.0f} KB')
